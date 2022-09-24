
import json
from openpyxl import load_workbook
import firebase_admin
from firebase_admin import credentials, storage
from flask import Flask, request, jsonify



def routine_adaptation(path):


    cred = credentials.Certificate("D:\\U can make\\routine\\lib\\api\key.json")
    firebase = firebase_admin.initialize_app(cred, {"storageBucket" : "routine-24ca0.appspot.com"})

    #path='routine2_040302.xlsx'

    path='files/'+path

    _bucket= storage.bucket()
    _file=_bucket.blob(path)
    _file.download_to_filename('current_file.xlsx')

   
    
    #book = load_workbook("D:\\Programs Files\\excel integration\\routine2.xlsx")
    #book = load_workbook(r"D:\ANIRBAN\Garbaze\routine2_040302.xlsx")
    book = load_workbook('current_file.xlsx')
    sheet = book.active
    
    #add the coordinate with every non-None cell.value
    for row in sheet:
        for cell in row:
            temp=str(cell.value)
            _coordinate=str(cell.coordinate)
            if cell.value !=None:
                sheet[cell.coordinate]=  temp+'*'+_coordinate     
    
    fullSheet=[]
    
    #making simple list of merged cells
    mergedCellStr = str(sheet.merged_cells)
    def mergedCellList(x):
        mergedCell=[]
        _str=''
        i=0
        while i < len(x):
            
            if x[i]==' ':
                mergedCell+=[_str]
                _str=''
            else:
                _str+=x[i]
        
            i+=1
        mergedCell+=[_str]
        return mergedCell
    
    _mergedCells=mergedCellList(mergedCellStr)
    
    
    #print(_mergedCells)
    
    #removing all_none valued row
    for row in sheet:
         tempList = [cell.value for cell in  row]
         if tempList != [None]*len(tempList):  
                fullSheet += [tempList]
    
                       
            
    fullReadableSheet=[]
    for row in fullSheet:  # removing non important row
        if row.count(None)<=len(row)-7:
            fullReadableSheet+=[row]
    
    
    firstRow=fullReadableSheet[0]
    
    i=0
    while firstRow[i]==None: #counting all_None column 
        i+=1
    
    for row in fullReadableSheet: # removing all_none column
        for j in range(i):
            row.remove(None)
    
    
    #initialize day
    saturDay=[]
    sunDay=[]
    monDay=[]
    tuesDay=[]
    wednesDay=[]
    
    #setting header
    for row in fullReadableSheet:
        saturDay+=[[]]
        sunDay+=[[]]
        monDay+=[[]]
        tuesDay+=[[]]
        wednesDay+=[[]]
    #print(len(fullSheet2))
    
    #removing coordinate from cell value
    def realName(element):
        starPosition=element.find('*')
        element=element[0:starPosition]
        return element
    
    #class name
    classNames=[]
    item=2
    while item<len(fullReadableSheet):
        classNames+=[
            realName(fullReadableSheet[item][0])+' '+
            realName(fullReadableSheet[item+1][0])+' '+
            realName(fullReadableSheet[item+2][0])
            ]
        item+=3
    
    #print(classNames)
    #assigning value from fullSheet2 in each row
     
    for i in range(len(fullReadableSheet)):
        for j in range(1,10):
            saturDay[i]+=[fullReadableSheet[i][j]]
        for j in range(10,19):
            sunDay[i]+=[fullReadableSheet[i][j]]
        for j in range(19,28):
            monDay[i]+=[fullReadableSheet[i][j]]
        for j in range(28,37):
            tuesDay[i]+=[fullReadableSheet[i][j]]
        for j in range(37,46):
            wednesDay[i]+=[fullReadableSheet[i][j]]    
    
    #for row in saturDay:
        #print(row)
    #print(saturDay)
    
    
    ## Dictionary/Json data of the routine 
    
    fullOrganizedClass={}
    
    #initializing fullOrganizedClass
    for item in classNames:
        fullOrganizedClass[item]={'sat':[],'sun':[],'mon':[],'tue':[],'wed':[]}
    #print(fullOrganizedClass)
    
    
    #extracting cell coordinate from cell value
    def cellCoordinate (element):
        starPosition=element.find('*')
        element=element[starPosition+1:]
        return element
    
    #finding the length of each class
    def cellLenght (element):
        element=cellCoordinate(element)
        for i in _mergedCells:
            colon= i.find(':')
            temp=i[:colon]
            if element == temp:
                element=i
        x1=element[:element.find(':')]
        x2=element[element.find(':')+1:]
        def stringAsciiValue(str):
                sum=0
                count=0
                for i in str:
                        if ord(i)>ord('9'):
                                sum+=ord(i)-ord('A')
                                count+=1
                return sum+count*26
        return stringAsciiValue(x2)-stringAsciiValue(x1)+1
    
    #inserting the value into fullOrganized class
    def classOrganizer(toDay,today_name):
        classes={}
        i=2
        k=0
        while i<len(toDay):
            className=classNames[k]
            classes[className]=''
            
            j=0;
            classSet=[]
            while j<len(toDay[i]):
                if(toDay[i][j]!=None):
                    #print(saturDay[i][j])
                    _class={}
                    classTime=toDay[1][j]
                    _classTime=''
                    ii=0
                    while classTime[ii]!='-':
                        _classTime+=classTime[ii]
                        ii+=1
                    ii=-1
                    while _classTime[ii]==' ':
                        _classTime=_classTime[:-1]
                    _class['Class Time']=_classTime
                    _class['Class Name']= realName(str(toDay[i][j]))
                    _class['Teacher Name']=realName(str(toDay[i+1][j]))
                    _class['Room No']=realName(str(toDay[i+2][j]))
                    _class['Class Lenght']=min(cellLenght(str(toDay[i][j])),cellLenght(str(toDay[i+2][j])))
                    classSet+=[_class]
    
                
                j+=1
    
            classes[className]=classSet
                
            k+=1
            i+=3
        
        #print(classes)
        for item in classes:
            fullOrganizedClass[item][today_name]=classes[item]
    
    classOrganizer(saturDay,'sat')
    classOrganizer(sunDay,'sun')
    classOrganizer(monDay,'mon')
    classOrganizer(tuesDay,'tue')
    classOrganizer(wednesDay,'wed')
    
    #printing full dictionary
    for item in fullOrganizedClass:
        print(item+":")
        temp=fullOrganizedClass[item]
        for i in temp:
            print(i+':')
            for j in temp[i]:
                print(j)
    return fullOrganizedClass




app = Flask(__name__)  
 
# A decorator used to tell the application
# which URL is associated function
@app.route('/user', methods =["GET", "POST"])


def show():
  
    global response
    if request.method == "POST":
       # getting input with name = fname in HTML form
       request_data=request.data
       data=json.loads(request_data.decode('utf-8'))
       
       print(data)
       print(data['excel'])
       path=data['excel']
       response=routine_adaptation(path)
       return " "
    else:
       return jsonify({'finalPath': response})
    
   
    
 
if __name__=='__main__':
   app.run()